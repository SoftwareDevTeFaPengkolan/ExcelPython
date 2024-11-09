from openpyxl import load_workbook

wb_row = load_workbook('hello_world.xlsx')


#Accessing sheet by name

sheet = wb_row["Sheet"]

#Print last row number by max_row attribute

print("Last Row Number: ", sheet.max_row)